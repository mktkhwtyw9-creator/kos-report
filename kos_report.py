#!/usr/bin/env python3
"""KOS聚光投流时报 — GitHub Actions 云端版

完整链路：IMAP收件 → 解析xlsx → 生成BI图片 → 推git托管 → 发钉钉webhook
全程在GitHub Actions云端运行，不依赖本地电脑。

环境变量：
  IMAP_SERVER   IMAP服务器地址
  IMAP_PORT     IMAP端口
  IMAP_USER     邮箱账号
  IMAP_PASSWORD 邮箱授权码
  CLIENT_WEBHOOK 钉钉客户群webhook URL
  GITHUB_TOKEN   GitHub Actions自动提供
  GITHUB_REPOSITORY GitHub Actions自动提供 (格式: username/repo)
"""
import json
import os
import sys
import imaplib
import socket
import email
import subprocess
import time
from email.header import decode_header
from datetime import datetime, timezone, timedelta

socket.setdefaulttimeout(30)

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow", "-q"])
    from PIL import Image, ImageDraw, ImageFont

try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "-q"])
    import requests

# ============== 配置（从环境变量读取） ==============
IMAP_SERVER = os.environ.get("IMAP_SERVER", "imap.exmail.qq.com")
IMAP_PORT = int(os.environ.get("IMAP_PORT", "993"))
IMAP_USER = os.environ.get("IMAP_USER", "can.yang@tarsocial.com")
# 兼容 workflow 中的 IMAP_PASS 和脚本的 IMAP_PASSWORD
IMAP_PASSWORD = os.environ.get("IMAP_PASSWORD") or os.environ.get("IMAP_PASS", "")
SENDER_FILTER = "service@mobgi.com"
# 兼容 workflow 中的 DINGTALK_WEBHOOK 和脚本的 CLIENT_WEBHOOK
CLIENT_WEBHOOK = os.environ.get("CLIENT_WEBHOOK") or os.environ.get("DINGTALK_WEBHOOK", "")
KEYWORD = "创量"
GITHUB_REPOSITORY = os.environ.get("GITHUB_REPOSITORY", "")
# 兼容 workflow 中的 GIT_TOKEN（PAT）和 GitHub Actions 自动提供的 GITHUB_TOKEN
GITHUB_TOKEN = os.environ.get("GIT_TOKEN") or os.environ.get("GITHUB_TOKEN", "")

# 北京时间
BJT = timezone(timedelta(hours=8))

# 图片输出目录（相对于repo根目录）
IMAGE_DIR = "images"


# ============== IMAP 收件 ==============

def connect_imap():
    """连接IMAP服务器"""
    mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail.login(IMAP_USER, IMAP_PASSWORD)
    return mail


def decode_mime_header(header_val):
    if not header_val:
        return ""
    parts = decode_header(header_val)
    result = []
    for part, charset in parts:
        if isinstance(part, bytes):
            try:
                result.append(part.decode(charset or "utf-8", errors="replace"))
            except Exception:
                result.append(part.decode("utf-8", errors="replace"))
        else:
            result.append(str(part))
    return "".join(result)


def find_chuangliang_reports(mail, date_str=None):
    """搜索当日创量报表邮件"""
    if date_str is None:
        now = datetime.now(BJT)
        date_str = now.strftime("%d-%b-%Y")

    mail.select("INBOX")
    result, email_ids = mail.search(None, f'(SINCE {date_str})')
    if result != "OK" or not email_ids[0]:
        return []

    email_id_list = email_ids[0].split()
    email_id_list = email_id_list[-50:] if len(email_id_list) > 50 else email_id_list

    print(f"  当日邮件总数: {len(email_ids[0].split())}, 取最近: {len(email_id_list)}")
    reports = []

    for eid in email_id_list:
        try:
            result, msg_data = mail.fetch(eid, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT DATE)])")
            if result != "OK":
                continue
            raw_header = msg_data[0][1]
            msg_hdr = email.message_from_bytes(raw_header)
            subject = decode_mime_header(msg_hdr.get("Subject", ""))
            from_addr = decode_mime_header(msg_hdr.get("From", ""))
        except Exception as e:
            print(f"  获取邮件头失败 (ID:{eid}): {e}", file=sys.stderr)
            continue

        if SENDER_FILTER not in from_addr:
            continue

        print(f"  命中创量邮件 ID:{eid.decode()} 主题:{subject[:50]}")

        if "测试消息" in subject and "报表" not in subject:
            print(f"    -> 跳过测试消息")
            continue

        try:
            result, msg_data = mail.fetch(eid, "(RFC822)")
            if result != "OK":
                continue
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
        except Exception as e:
            print(f"  获取邮件正文失败 (ID:{eid}): {e}", file=sys.stderr)
            continue

        attachments = []
        for part in msg.walk():
            content_disposition = str(part.get("Content-Disposition", ""))
            if "attachment" in content_disposition:
                filename = part.get_filename()
                if filename:
                    filename = decode_mime_header(filename)
                    if filename.lower().endswith(".xlsx"):
                        attachments.append({
                            "filename": filename,
                            "data": part.get_payload(decode=True),
                        })

        if attachments:
            reports.append({
                "uid": eid.decode(),
                "subject": subject,
                "from": from_addr,
                "date": msg.get("Date", ""),
                "attachments": attachments,
            })

    return reports


# ============== xlsx 解析 ==============

def parse_xlsx_data(xlsx_bytes, filename=""):
    """解析创量xlsx报表"""
    import io
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    best_sheet = None
    best_rows = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        if len(non_empty) > len(best_rows):
            best_rows = non_empty
            best_sheet = sheet_name

    if not best_rows:
        return [], [], None

    header_row_idx = None
    header_fields = ["消耗", "账号", "阅读", "CPM", "CPC", "CPE", "账户", "广告", "日期", "时间"]
    for i, row in enumerate(best_rows):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        matched = sum(1 for f in header_fields if any(f in s for s in row_strs))
        if matched >= 3:
            header_row_idx = i
            break

    if header_row_idx is None:
        header_row_idx = 0

    raw_headers = [str(c).strip() if c is not None else "" for c in best_rows[header_row_idx]]
    data_rows = best_rows[header_row_idx + 1:]

    COST_KEYS = ["消耗", "花费", "投放花费", "现金消耗", "总消耗"]
    READ_KEYS = ["阅读", "阅读量", "展现", "展现量", "曝光", "曝光量"]
    INTERACT_KEYS = ["互动", "互动量", "互动数", "总互动"]
    CPM_KEYS = ["CPM", "阅读CPM", "阅读cpm", "千次展现成本", "千次曝光", "千次阅读"]
    CPC_KEYS = ["CPC", "单次点击成本", "点击成本"]
    CPE_KEYS = ["CPE", "单次互动成本", "互动成本"]
    ACCOUNT_KEYS = ["账号", "账号名称", "账户名称", "账户"]
    DATE_KEYS = ["日期"]
    TIME_KEYS = ["时间"]

    def find_col(row, keys):
        for i, h in enumerate(raw_headers):
            h_lower = h.lower()
            for k in keys:
                if k.lower() in h_lower:
                    return i
        return -1

    col_cost = find_col(raw_headers, COST_KEYS)
    col_read = find_col(raw_headers, READ_KEYS)
    col_interact = find_col(raw_headers, INTERACT_KEYS)
    col_cpm = find_col(raw_headers, CPM_KEYS)
    col_cpc = find_col(raw_headers, CPC_KEYS)
    col_cpe = find_col(raw_headers, CPE_KEYS)
    col_account = find_col(raw_headers, ACCOUNT_KEYS)
    col_date = find_col(raw_headers, DATE_KEYS)
    col_time = find_col(raw_headers, TIME_KEYS)

    if col_cost < 0 or col_read < 0:
        if len(raw_headers) >= 6:
            col_date = 0
            col_time = 1
            col_account = 2 if len(raw_headers) > 2 else 0
            col_cost = 3
            col_read = 4
            col_interact = 5
            col_cpm = 6 if len(raw_headers) > 6 else -1
            col_cpc = 7 if len(raw_headers) > 7 else -1
            col_cpe = 8 if len(raw_headers) > 8 else -1

    if col_cost < 0:
        print(f"无法识别消耗列，表头: {raw_headers}", file=sys.stderr)
        return [], [], None

    def safe_val(row, col_idx):
        if col_idx < 0 or col_idx >= len(row):
            return ""
        v = row[col_idx]
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y/%m/%d %H:%M")
        if isinstance(v, float):
            s = f"{v:.4f}".rstrip("0").rstrip(".")
            return s
        return str(v).strip()

    # 提取日期+时间
    date_str = None
    if col_date >= 0:
        for row in data_rows:
            v = safe_val(row, col_date)
            if v:
                for fmt in ["%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%Y年%m月%d日 %H:%M", "%Y年%m月%d日"]:
                    try:
                        dt = datetime.strptime(v.strip(), fmt)
                        date_str = dt.strftime("%Y/%m/%d %H:%M")
                        break
                    except ValueError:
                        continue
                if date_str:
                    break

    if col_time >= 0:
        for row in data_rows:
            t = safe_val(row, col_time)
            if t:
                t_clean = t.strip()
                for fmt_t in ["%H:%M", "%H:%M:%S", "%H时%M分"]:
                    try:
                        tm = datetime.strptime(t_clean, fmt_t)
                        time_part = tm.strftime("%H:%M")
                        if date_str:
                            date_part = date_str.split(" ")[0]
                            date_str = f"{date_part} {time_part}"
                        else:
                            date_str = f"{datetime.now(BJT).strftime('%Y/%m/%d')} {time_part}"
                        break
                    except ValueError:
                        continue
                if date_str and ":" in date_str.split(" ")[1]:
                    break

    if not date_str:
        date_str = datetime.now(BJT).strftime("%Y/%m/%d %H:%M")

    out_headers = ["日期", "时间", "账号名称", "消耗", "阅读", "互动", "阅读CPM", "CPC", "CPE"]

    formatted_rows = []
    for row in data_rows:
        account = safe_val(row, col_account) if col_account >= 0 else ""
        cost = safe_val(row, col_cost)

        if not account and not cost:
            continue
        if account in ("合计", "总计", "代理合计", "") and not cost:
            continue
        if cost in ("-", "", "0", "0.0"):
            continue

        date_part = date_str.split(" ")[0] if " " in date_str else date_str
        time_part = date_str.split(" ")[1] if " " in date_str else ""

        row_data = [
            date_part,
            time_part,
            account,
            cost,
            safe_val(row, col_read),
            safe_val(row, col_interact),
            safe_val(row, col_cpm),
            safe_val(row, col_cpc),
            safe_val(row, col_cpe),
        ]
        formatted_rows.append(row_data)

    return out_headers, formatted_rows, date_str


# ============== 图片生成 ==============

def _get_font(size):
    """获取中文字体（兼容Linux和Windows）"""
    font_paths = [
        # Linux (GitHub Actions / Ubuntu)
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansSC-Regular.otf",
        "/usr/share/fonts/opentype/noto/NotoSansSC-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        # Windows
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/msyh.ttf",
        "C:/Windows/Fonts/msyhbd.ttc",
        "C:/Windows/Fonts/simhei.ttf",
    ]
    for p in font_paths:
        if os.path.exists(p):
            return ImageFont.truetype(p, size)
    # 最后尝试加载任何可用字体
    raise RuntimeError("未找到中文字体")


def generate_report_image(rows, date_time, output_path=None):
    """生成BI卡片风格的表格图片"""
    headers = ["日期", "时间", "账号名称", "消耗", "阅读", "互动", "阅读CPM", "CPC", "CPE"]

    title_font = _get_font(26)
    subtitle_font = _get_font(16)
    header_font = _get_font(17)
    cell_font = _get_font(15)

    padding = 14
    min_col_width = 60

    tmp_img = Image.new("RGB", (1, 1), "white")
    draw = ImageDraw.Draw(tmp_img)

    def text_width(text, font):
        bbox = draw.textbbox((0, 0), str(text), font=font)
        return bbox[2] - bbox[0]

    col_widths = []
    for i, h in enumerate(headers):
        w = text_width(h, header_font) + padding * 2
        for row in rows:
            if i < len(row):
                cell_w = text_width(str(row[i]), cell_font) + padding * 2
                w = max(w, cell_w)
        w = max(w, min_col_width)
        col_widths.append(w)

    table_width = sum(col_widths) + len(headers) + 1
    row_height = 40
    header_height = 44
    title_height = 70
    margin = 28
    footer_height = 30

    img_width = table_width + margin * 2
    img_height = title_height + header_height + len(rows) * row_height + margin * 2 + footer_height

    img = Image.new("RGB", (img_width, img_height), "white")
    draw = ImageDraw.Draw(img)

    yellow = "#FFD54F"
    yellow_dark = "#F5B400"
    dark = "#222222"
    gray = "#777777"
    light_gray = "#FAFAFA"
    border_color = "#DDDDDD"

    # 标题
    title_text = "创量 KOS聚光投放时报（当日累计）"
    bbox = draw.textbbox((0, 0), title_text, font=title_font)
    title_w = bbox[2] - bbox[0]
    draw.text(((img_width - title_w) / 2, margin), title_text, font=title_font, fill=dark)
    draw.line([(img_width / 2 - title_w / 2, margin + 35),
               (img_width / 2 + title_w / 2, margin + 35)], fill=yellow_dark, width=3)

    # 副标题
    bbox = draw.textbbox((0, 0), date_time, font=subtitle_font)
    sub_w = bbox[2] - bbox[0]
    draw.text(((img_width - sub_w) / 2, margin + 42), date_time, font=subtitle_font, fill=gray)

    table_x = margin
    table_y = margin + title_height

    # 表头
    x = table_x
    for i, w in enumerate(col_widths):
        draw.rectangle([x, table_y, x + w, table_y + header_height], fill=yellow)
        x += w

    x = table_x
    for i, h in enumerate(headers):
        w = col_widths[i]
        bbox = draw.textbbox((0, 0), h, font=header_font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        tx = x + (w - text_w) / 2
        ty = table_y + (header_height - text_h) / 2
        draw.text((tx, ty), h, font=header_font, fill=dark)
        x += w

    # 数据行
    y = table_y + header_height
    for row_idx, row in enumerate(rows):
        bg = light_gray if row_idx % 2 == 1 else "white"
        x = table_x
        for i, cell in enumerate(row):
            w = col_widths[i]
            draw.rectangle([x, y, x + w, y + row_height], fill=bg)
            bbox = draw.textbbox((0, 0), str(cell), font=cell_font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
            tx = x + (w - text_w) / 2
            cell_y = y + (row_height - text_h) / 2
            draw.text((tx, cell_y), str(cell), font=cell_font, fill=dark)
            x += w
        y += row_height

    # 表格边框
    x = table_x
    for i in range(len(headers) + 1):
        draw.line([(x, table_y),
                   (x, table_y + header_height + len(rows) * row_height)],
                  fill=border_color, width=1)
        x += col_widths[i] if i < len(col_widths) else 0

    y = table_y
    for i in range(len(rows) + 2):
        draw.line([(table_x, y), (table_x + table_width, y)],
                  fill=border_color, width=1)
        y += header_height if i == 0 else row_height

    # 页脚
    footer = f"数据由{KEYWORD}机器人自动采集 · 仟传传播"
    bbox = draw.textbbox((0, 0), footer, font=subtitle_font)
    foot_w = bbox[2] - bbox[0]
    draw.text(((img_width - foot_w) / 2, img_height - footer_height - 4),
              footer, font=subtitle_font, fill=gray)

    if output_path:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        img.save(output_path, "PNG", optimize=True)
    return img, img_width, img_height


# ============== Git 推送图片 ==============

def git_push_image(image_path, image_filename):
    """将图片提交并推送到GitHub仓库"""
    repo = GITHUB_REPOSITORY
    token = GITHUB_TOKEN

    if not repo or not token:
        print("GITHUB_REPOSITORY 或 GITHUB_TOKEN 未设置，跳过git推送")
        return None

    # 配置远程URL带token认证
    auth_url = f"https://x-access-token:{token}@github.com/{repo}.git"
    
    # 清除 actions/checkout 设置的 extraheader（它会用自动GITHUB_TOKEN覆盖PAT）
    subprocess.run(["git", "config", "--local", "--unset-all", "http.https://github.com/.extraheader"], capture_output=True)
    
    subprocess.run(["git", "remote", "set-url", "origin", auth_url], check=True)

    # 配置用户信息
    subprocess.run(["git", "config", "user.name", "github-actions[bot]"], check=True)
    subprocess.run(["git", "config", "user.email", "github-actions[bot]@users.noreply.github.com"], check=True)

    # 添加图片
    subprocess.run(["git", "add", image_path], check=True)

    # 提交
    commit_msg = f"auto: add report image {image_filename}"
    result = subprocess.run(["git", "commit", "-m", commit_msg], capture_output=True, text=True)
    if result.returncode != 0:
        print(f"git commit: {result.stdout} {result.stderr}")
        if "nothing to commit" in result.stdout:
            return None
        # 非致命错误，继续

    # 推送
    result = subprocess.run(["git", "push"], capture_output=True, text=True)
    if result.returncode != 0:
        print(f"git push failed: {result.stderr}", file=sys.stderr)
        return None

    print(f"图片已推送到GitHub: {image_filename}")
    return True


def get_image_url(image_filename, repo=None):
    """构造图片的公网URL（jsDelivr CDN，国内可访问）"""
    if not repo:
        repo = GITHUB_REPOSITORY
    # jsDelivr CDN URL（有国内镜像，速度快）
    return f"https://cdn.jsdelivr.net/gh/{repo}@main/{IMAGE_DIR}/{image_filename}"


# ============== 钉钉推送 ==============

def send_dingtalk_webhook(rows, date_time, pic_url):
    """发送markdown消息到钉钉客户群"""
    # 取第一行数据
    row = rows[0]
    date_part, time_part, account, cost, read, interact, cpm, cpc, cpe = row

    # 表格行（取前3行）
    table_rows = "\n".join(
        f"| {r[3]} | {r[4]} | {r[5]} | {r[6]} | {r[7]} | {r[8]} |"
        for r in rows[:3]
    )

    markdown_text = f"""# 创量 KOS聚光投放时报（当日累计）

**时间**：{date_time}
**账号**：{account}

| 消耗 | 阅读 | 互动 | 阅读CPM | CPC | CPE |
| --- | --- | --- | --- | --- | --- |
{table_rows}

![创量 KOS聚光投放时报]({pic_url})

数据由创量机器人自动采集 · 仟传传播"""

    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": f"创量 KOS聚光投放时报 {date_time}",
            "text": markdown_text
        }
    }

    resp = requests.post(CLIENT_WEBHOOK, json=payload, timeout=15)
    result = resp.json()
    print(f"钉钉推送结果: {result}")
    return result


# ============== 主流程 ==============

def main(retry=0, max_retries=5, retry_delay=30):
    """主流程：收邮件 → 解析 → 生成图片 → 推git → 发webhook"""
    now = datetime.now(BJT)
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] KOS时报云端推送 启动 (重试 {retry}/{max_retries})")

    # 1. 连接IMAP
    print("连接IMAP...")
    try:
        mail = connect_imap()
    except Exception as e:
        print(f"IMAP连接失败: {e}", file=sys.stderr)
        return False

    date_str = now.strftime("%d-%b-%Y")
    print(f"搜索当日({date_str})创量邮件...")
    reports = find_chuangliang_reports(mail, date_str)
    mail.logout()

    if not reports:
        print("未找到创量xlsx报表邮件")
        if retry < max_retries:
            print(f"等待 {retry_delay} 秒后重试 ({retry+1}/{max_retries})...")
            time.sleep(retry_delay)
            return main(retry=retry+1, max_retries=max_retries, retry_delay=retry_delay)
        print("已达最大重试次数，放弃")
        return False

    print(f"找到 {len(reports)} 封创量报表邮件")

    # 2. 解析最新邮件
    latest = reports[-1]
    print(f"处理邮件: {latest['subject']} ({len(latest['attachments'])} 个xlsx附件)")

    all_rows = []
    date_str = None

    for att in latest["attachments"]:
        print(f"  解析附件: {att['filename']}")
        try:
            headers, rows, ds = parse_xlsx_data(att["data"], att["filename"])
            if rows:
                all_rows.extend(rows)
                if ds and not date_str:
                    date_str = ds
        except Exception as e:
            print(f"  解析失败: {e}", file=sys.stderr)

    if not all_rows:
        print("所有附件均无可解析的数据行")
        return False

    if not date_str:
        date_str = now.strftime("%Y/%m/%d %H:%M")

    # 3. 生成图片
    print(f"生成BI表格图片 (日期: {date_str}, 数据行: {len(all_rows)})")
    ts = int(time.time())
    image_filename = f"report-{ts}.png"
    image_path = os.path.join(IMAGE_DIR, image_filename)

    img, w, h = generate_report_image(all_rows, date_str, image_path)
    print(f"图片已保存: {image_path} ({w}x{h})")

    # 数据预览
    print(f"\n数据预览 ({len(all_rows)} 行):")
    for i, row in enumerate(all_rows[:3]):
        print(f"  [{i+1}] 账号: {row[2]}, 消耗: {row[3]}, 阅读: {row[4]}, 互动: {row[5]}, CPM: {row[6]}, CPC: {row[7]}, CPE: {row[8]}")

    # 4. 推送图片到GitHub
    print("\n推送图片到GitHub...")
    push_ok = git_push_image(image_path, image_filename)

    if not push_ok:
        print("图片推送失败，无法获取公网URL")
        return False

    # 5. 构造图片URL并推送钉钉
    pic_url = get_image_url(image_filename)
    print(f"图片公网URL: {pic_url}")

    # 等待几秒让CDN同步
    print("等待CDN同步（10秒）...")
    time.sleep(10)

    print("\n推送钉钉客户群...")
    result = send_dingtalk_webhook(all_rows, date_str, pic_url)

    if result.get("errcode") == 0:
        print("\n✅ 推送成功！")
        return True
    else:
        print(f"\n❌ 推送失败: {result}")
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
